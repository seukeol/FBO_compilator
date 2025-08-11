from aiohttp import ClientSession
import asyncio
import json


CLIENT_ID = '1'
API_TOKEN = '1'
HEADERS = {
    "Client-Id": CLIENT_ID,
    "Api-Key": API_TOKEN,
    "Content-Type": "application/json"
}


data = {
    'order_ids': [
        input('Введите номер поставки:')
        ]
}


final_data = {}
names_dict = {}

async def names():
    async with ClientSession() as session:
        dataaa = {
            "cluster_ids": [],
            "cluster_type": "CLUSTER_TYPE_OZON"
        }
        async with session.post('https://api-seller.ozon.ru/v1/cluster/list',data=json.dumps(dataaa),headers=HEADERS) as response:
            resp_json = await response.json()
            for cluster in resp_json['clusters']:
                for logistic_cluster in cluster['logistic_clusters']:
                    for sklad in logistic_cluster['warehouses']:
                        names_dict[sklad['warehouse_id']] = sklad['name']


async def find_name(id):
    if not names_dict:
        await names()
    try:
        return names_dict[id]
    except KeyError:
        return id


async def main():
    async with ClientSession() as session:
        async with session.post('https://api-seller.ozon.ru/v2/supply-order/get', data = json.dumps(data),headers=HEADERS) as response:
            print("Status:", response.status)
            print("Content-type:", response.headers['content-type'])
            resp_json=await response.json()
            print(resp_json['orders'])




        for i in resp_json['orders'][0]['supplies']:
            name = await find_name(i['storage_warehouse_id'])
            bundle_data = {
                "bundle_ids": [
                    i['bundle_id']
                ],
                "is_asc": True,
                "storage_warehouse_ids": [i['storage_warehouse_id']],
                "dropoff_warehouse_id": resp_json['orders'][0]['dropoff_warehouse_id'],
                "limit": 100
            }
            print(i['storage_warehouse_id'])
            print(name)
            await asyncio.sleep(1)

            async with session.post('https://api-seller.ozon.ru/v1/supply-order/bundle', data=json.dumps(bundle_data),headers=HEADERS) as response:
                print("Status:", response.status)
                bundles_resp = await response.json()

                for j in bundles_resp['items']:
                    try:
                        final_data[name].append([j['offer_id'],j['quantity']])
                    except:
                        final_data[name] = [[j['offer_id'],j['quantity']]]






asyncio.run(main())


for i in final_data.values():
    for j in range(len(i)):
        i[j].append('')
        i[j][1], i[j][2] = i[j][2], i[j][1]

print(final_data)

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
counter = 1


for i in final_data.keys():
    ws[f'A{counter}'] = str(i)
    ws.merge_cells(start_row=counter, end_row=counter, start_column=1, end_column=3)
    counter +=1
    for j in final_data[i]:
        ws.append(j)
        counter += 1
    counter+=2
wb.save("sample.xlsx")

print('seu<3')